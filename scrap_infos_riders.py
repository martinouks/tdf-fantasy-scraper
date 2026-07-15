import os
import sys
import requests
from pathlib import Path
from typing import List, Dict, Any
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.formatting.rule import FormulaRule
from dotenv import load_dotenv

load_dotenv()

DIR_SHEETS_FILES = "spreadsheets"

API_URL = "https://fantasybytissot.letour.fr/v1/private/searchjoueurs?lg=fr"

# adapt the token in .env file
# adpat X-Access-Key with the one in the network request headers
HEADERS = {
    'User-Agent': 'Mozilla/5.0 (X11; Ubuntu; Linux x86_64; rv:152.0) Gecko/20100101 Firefox/152.0',
    'Accept': 'application/json',
    'Content-Type': 'application/json',
    'X-Access-Key': '630@20.31@',
    'Authorization': os.getenv("AUTH_TOKEN"),
    'Origin': 'https://fantasybytissot.letour.fr',
    'Referer': 'https://fantasybytissot.letour.fr/',
}

# adapt idj with the current stage
PAYLOAD = {
    "filters": {
        "nom": "", "club": "", "position": "", "budget_ok": False,
        "valeur_max": 27, "engage": False, "partant": False, 
        "dreamteam": False, "quota": "", "idj": "12", 
        "pageIndex": 0, "pageSize": 1000, 
        "loadSelect": 1, "searchonly": 1
    }
}

COLORS = {
    "leader": "1fa000",
    "rouleur": "ffba00",
    "grimpeur": "ff7800",
    "baroudeur": "0094e1",
    "Unknown": "FFFFFF",
    "Blue_sky": "deebf7",
    "Yellow": "FFFF00"
}

EXCEL_HEADERS = ["NOM", "Equipe", "Coût", "Nb Offres", "Infos sup", "Etoiles misées", "Appartient à", "", "Étoiles restantes", "=200-SUM(F2:F500)", "Coureurs", "=COUNT(F2:F500)"]
WIDTHS_COLS = {'A': 35, 'B': 30, 'C': 10, 'D': 10, 'E': 40, 'F': 15, 'G': 20, 'H': 10, 'I': 35, 'J': 10, 'K': 20, 'L': 10}

LETTER_TO_INDEX = {chr(i + 64): i for i in range(1, 27)}

def _fetch_data_api() -> List[Dict[str, Any]]:
    """Fetches data from the API and store it a list of dictionaries."""
    print("Fetching live data from Fantasy API...")

    response = requests.post(API_URL, headers=HEADERS, json=PAYLOAD)
    if response.status_code != 200:
        print(f" API Error: {response.status_code} - {response.text}")
        return []

    raw_list = response.json().get('joueurs', [])
    riders = []

    for item in raw_list:
        riders.append({
            "name": item.get('nomcomplet', 'Unknown').strip(),
            "team": item.get('club', 'Unknown').strip(),
            "cost": int(item.get('valeur', 0)),
            "category": item.get('position', 'Unknown').strip().replace('lib_', ''),
            "current_offers": int(item.get('offres_encours_nb', 0)),
            "proprietaire": item.get('proprietaire', {id: ''})
        })

    print(f"Downloaded {len(riders)} active riders.")
    return riders

def _get_saved_manual_data(filepath: str) -> tuple:
    rider_notes = {}
    canvas_notes = {}
    
    if not Path(filepath).exists():
        print("No existing file found. A new one will be created from scratch.")
        return rider_notes, canvas_notes

    print(f"Reading existing file '{filepath}' to preserve current infos...")
    try:
        wb = load_workbook(filepath)
        ws = wb.active
        max_col = ws.max_column
        max_row = ws.max_row
        
        # Notes attached to riders
        for row in ws.iter_rows(min_row=2):
            rider_name = row[0].value
            if rider_name: 
                rider_notes[rider_name] = {
                    "E": row[LETTER_TO_INDEX["E"] - 1].value,
                    "F": row[LETTER_TO_INDEX["F"] - 1].value,
                    "H": row[LETTER_TO_INDEX["H"] - 1].value
                }

        # Canvas notes  (Cols J+)
        for r_idx in range(2, max_row + 1):
            for c_idx in range(LETTER_TO_INDEX["J"], max_col + 1):
                cell_val = ws.cell(row=r_idx, column=c_idx).value
                if cell_val is not None:
                    canvas_notes[(r_idx, c_idx)] = cell_val

    except Exception as e:
        print(f"Could not read existing data. Proceeding without merging. Error: {e}")

    return rider_notes, canvas_notes

def _format_cell(cell, fill_color: str = None, is_header: bool = False):
    """Apply styling to a cell."""
    if is_header:
        border = Border(bottom=Side(style='medium'))
    else:
        border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                        top=Side(style='thin'), bottom=Side(style='thin'))
    cell.border = border

    cell.alignment = Alignment(horizontal="center")

    if is_header:
        cell.font = Font(bold=True)
    if fill_color:
        cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")

def _export_to_excel(riders: List[Dict[str, Any]], saved_data: tuple, filepath: str):
    """Generates the Excel file, merging fresh API data with saved manual inputs."""
    print("Generating updated Excel file...")

    rider_notes, canvas_notes = saved_data

    wb = Workbook()
    ws = wb.active
    ws.title = "Fantasy TDF Riders"

    # Write and format headers
    ws.append(EXCEL_HEADERS)
    for col_idx in range(LETTER_TO_INDEX["A"], LETTER_TO_INDEX["L"] + 1):
        _format_cell(ws.cell(row=1, column=col_idx), is_header=True)

    ws.cell(row=1, column=LETTER_TO_INDEX["I"]).fill = PatternFill(start_color=COLORS["Yellow"], end_color=COLORS["Yellow"], fill_type="solid")
    ws.cell(row=1, column=LETTER_TO_INDEX["K"]).fill = PatternFill(start_color=COLORS["Yellow"], end_color=COLORS["Yellow"], fill_type="solid")

    ws.cell(row=1, column=LETTER_TO_INDEX["L"]).border = Border(right=Side(style='medium'), bottom=Side(style='medium'))

    # Sort riders by Team (alphabetical) then Cost (descending)
    sorted_riders = sorted(riders, key=lambda x: (x['team'], -x['cost']))

    row_idx = 2
    current_team = ""

    for rider in sorted_riders:
        # Add visual separation between teams
        if current_team and current_team != rider['team']:
            row_idx += 1
        current_team = rider['team']

        # Load rider data into the sheet
        ws.cell(row=row_idx, column=LETTER_TO_INDEX["A"], value=rider['name'])
        ws.cell(row=row_idx, column=LETTER_TO_INDEX["B"], value=rider['team'])
        ws.cell(row=row_idx, column=LETTER_TO_INDEX["C"], value=rider['cost'])

        offers = rider['current_offers']
        ws.cell(row=row_idx, column=LETTER_TO_INDEX["D"], value=offers if offers > 0 else "")

        proprietaire = rider['proprietaire']
        ws.cell(row=row_idx, column=LETTER_TO_INDEX["G"], value=proprietaire["nom"] if proprietaire["id"] else "")

        ws.cell(row=row_idx, column=LETTER_TO_INDEX["I"], value=f'=IF(F{row_idx}>0,A{row_idx},"")') # print the rider name if chosed
        ws.cell(row=row_idx, column=LETTER_TO_INDEX["I"]).alignment = Alignment(horizontal="center")

        # Merge saved notes
        notes = rider_notes.get(rider['name'], {"E": "", "F": "", "H": ""})
        ws.cell(row=row_idx, column=LETTER_TO_INDEX["E"], value=notes.get("E", ""))
        ws.cell(row=row_idx, column=LETTER_TO_INDEX["F"], value=notes.get("F", ""))
        ws.cell(row=row_idx, column=LETTER_TO_INDEX["H"], value=notes.get("H", ""))

        # Apply Styling
        hex_color = COLORS.get(rider['category'], COLORS["Unknown"])
        for col_idx in range(LETTER_TO_INDEX["A"], LETTER_TO_INDEX["G"] + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            # Only color the first 3 columns as requested
            apply_color = hex_color if col_idx in [LETTER_TO_INDEX["A"], LETTER_TO_INDEX["B"], LETTER_TO_INDEX["C"]] else None
            _format_cell(cell, fill_color=apply_color)

        row_idx += 1

    for (r, c), val in canvas_notes.items():
        ws.cell(row=r, column=c, value=val)

    # Conditional Formatting for riders name added
    blue_fill = PatternFill(start_color=COLORS["Blue_sky"], end_color=COLORS["Blue_sky"], fill_type="solid")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                         top=Side(style='thin'), bottom=Side(style='thin'))

    rule = FormulaRule(formula=['F2>0'], fill=blue_fill, border=thin_border)
    
    ws.conditional_formatting.add('I2:I500', rule)

    # Adjust Column Widths
    for col_letter, width in WIDTHS_COLS.items():
        ws.column_dimensions[col_letter].width = width

    ws.freeze_panes = "A2"  # Freeze the header row

    # Add rules
    ws.cell(row=3, column=10, value="Qq règles :")
    ws.cell(row=4, column=10, value="- on peut faire une offre -1 par rapport au prix actuel du coureur")
    ws.cell(row=5, column=10, value="- achat en secret : 20 crédits")
    ws.cell(row=6, column=10, value="- enchère accélérée (réduire le temps de validation à moins de 2h), dispo 2j après la 1ère session d'enchères : 30 crédits")
    ws.cell(row=7, column=10, value="- vente coureur : on récupère les étoiles que ça nous a coûté + si il a couru dans l'équipe, on paye 50 crédits")
    ws.cell(row=8, column=10, value="- 5 coureurs polyvalents, 3 grimpeurs, 3 sprinters et 3 leaders max par équipe")
    ws.cell(row=9, column=10, value="- validations d'enchères chaque nuit à 4h15 et au début de chaque étape")

    wb.save(filepath)
    print(f"Success! File '{filepath}' is updated and ready.")

if __name__ == "__main__":
    if len(sys.argv) != 2:
        print("Usage: python scrap_infos_riders.py <output_excel_file>")
        sys.exit(1)
    else:
        Path(DIR_SHEETS_FILES).mkdir(parents=True, exist_ok=True) # create the dir spreadsheets if it doesn't exist
        FILE_NAME = DIR_SHEETS_FILES + "/" + sys.argv[1]

    live_riders = _fetch_data_api()

    if live_riders:
        # Extract notes from previous file (if it exists)
        saved_data = _get_saved_manual_data(FILE_NAME)

        # Build new file combining both
        _export_to_excel(live_riders, saved_data, FILE_NAME)
